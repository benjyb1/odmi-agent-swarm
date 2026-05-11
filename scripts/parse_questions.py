"""Parse the official ODMI questionnaire data into structured question objects.

Uses the 2025_odm_questionnaire_data.xlsx from data.europa.eu which has:
- Sheet 'questionnaire': question_id, dimension, indicator, question, response_scoring
- Sheet 'merged_responses': all country responses

Usage:
    uv run python scripts/parse_questions.py
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

# Map the spreadsheet's dimension names to our shorter labels
DIMENSION_MAP = {
    "policy_dimension": "Policy",
    "portal_dimension": "Portal",
    "quality_dimension": "Quality",
    "impact_dimension": "Impact",
}

DATA_FILE = Path(__file__).parent.parent / "data" / "questions" / "2025_odm_questionnaire_data.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "data" / "questions" / "odmi_2025_questions.json"


def parse_questions() -> list[dict]:
    """Parse the official ODMI questionnaire sheet."""
    if not DATA_FILE.exists():
        print(f"File not found: {DATA_FILE}")
        print("Download from: https://data.europa.eu/sites/default/files/2025-12/2025_odm_questionnaire_data.xlsx")
        sys.exit(1)

    wb = load_workbook(DATA_FILE, read_only=True)
    ws = wb["questionnaire"]

    questions = []
    headers = None

    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if i == 0:
            # Header row
            headers = [str(c).strip() if c else "" for c in row]
            continue

        cells = {headers[j]: row[j] for j in range(min(len(headers), len(row))) if headers[j]}

        question_id = str(cells.get("question_id", "")).strip()
        dimension_raw = str(cells.get("dimension", "")).strip()
        indicator = str(cells.get("indicator", "")).strip()
        question_text = str(cells.get("question", "")).strip()
        scoring = str(cells.get("response_scoring", "")).strip()

        if not question_id or not question_text:
            continue

        dimension = DIMENSION_MAP.get(dimension_raw, dimension_raw)

        questions.append({
            "question_id": question_id,
            "dimension": dimension,
            "indicator": indicator,
            "question_text": question_text,
            "response_scoring": scoring,
        })

    wb.close()
    return questions


def main():
    questions = parse_questions()

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(questions, indent=2, ensure_ascii=False))

    print(f"Parsed {len(questions)} questions:")
    for dim in ["Policy", "Portal", "Quality", "Impact"]:
        count = sum(1 for q in questions if q["dimension"] == dim)
        if count:
            print(f"  {dim}: {count}")

    # Show indicator breakdown
    indicators = {}
    for q in questions:
        ind = q["indicator"]
        indicators[ind] = indicators.get(ind, 0) + 1
    print(f"\nIndicators ({len(indicators)} total):")
    for ind, count in sorted(indicators.items()):
        print(f"  {ind}: {count}")

    print(f"\nOutput: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
