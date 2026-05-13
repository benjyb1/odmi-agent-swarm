"""Mirror ODMI's `merged_responses` sheet into the SQLite ground_truth table.

Per D22, ground truth replaces hand-marks as the evaluation set. The
xlsx is canonical; this script is the loader. Idempotent: re-runs
upsert by (question_id, country_code).

    uv run python scripts/load_ground_truth.py

Reads: data/questions/2025_odm_questionnaire_data.xlsx
       (sheet: merged_responses)
Writes: data/odmi.db (table: ground_truth)
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
DB_PATH = REPO_ROOT / "data" / "odmi.db"
XLSX_PATH = (
    REPO_ROOT / "data" / "questions" / "2025_odm_questionnaire_data.xlsx"
)


SCHEMA = """
CREATE TABLE IF NOT EXISTS ground_truth (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    question_id     TEXT NOT NULL,
    country_code    TEXT NOT NULL,
    country_name    TEXT,
    country_group   TEXT,
    cross_year_id   TEXT,
    dimension       TEXT,
    indicator       TEXT,
    response        TEXT,
    decision        TEXT,
    awarded_score   REAL,
    max_score       REAL,
    explanation     TEXT,
    cycle_year      INTEGER NOT NULL DEFAULT 2025,
    loaded_at       TEXT DEFAULT (datetime('now')),
    UNIQUE(question_id, country_code, cycle_year)
);

CREATE INDEX IF NOT EXISTS idx_gt_question_country
    ON ground_truth(question_id, country_code);
CREATE INDEX IF NOT EXISTS idx_gt_country
    ON ground_truth(country_code);
CREATE INDEX IF NOT EXISTS idx_gt_dimension
    ON ground_truth(dimension);
"""


COLUMN_ORDER = [
    "question_id", "country_code", "country_name", "country_group",
    "cross_year_id", "dimension", "indicator",
    "response", "decision", "awarded_score", "max_score", "explanation",
]


def _to_number(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main() -> None:
    if not XLSX_PATH.exists():
        raise SystemExit(f"Source xlsx missing: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH, read_only=True)
    if "merged_responses" not in wb.sheetnames:
        raise SystemExit("Sheet 'merged_responses' not found in xlsx.")

    ws = wb["merged_responses"]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c else "" for c in next(rows)]

    missing = [c for c in COLUMN_ORDER if c not in header]
    if missing:
        raise SystemExit(f"Expected xlsx columns missing: {missing}")

    payloads = []
    for r in rows:
        cells = {h: v for h, v in zip(header, r) if h}
        qid = cells.get("question_id")
        cc = cells.get("country_code")
        if not qid or not cc:
            continue
        payloads.append((
            str(qid).strip(),
            str(cc).strip(),
            cells.get("country_name"),
            cells.get("country_group"),
            cells.get("cross_year_id"),
            cells.get("dimension"),
            cells.get("indicator"),
            cells.get("response"),
            cells.get("decision"),
            _to_number(cells.get("awarded_score")),
            _to_number(cells.get("max_score")),
            cells.get("explanation"),
        ))
    wb.close()

    inserted = 0
    updated = 0
    with sqlite3.connect(DB_PATH) as conn:
        conn.executescript(SCHEMA)
        cur = conn.cursor()
        for p in payloads:
            qid, cc = p[0], p[1]
            existing = cur.execute(
                "SELECT id FROM ground_truth "
                "WHERE question_id = ? AND country_code = ? AND cycle_year = ?",
                (qid, cc, 2025),
            ).fetchone()
            if existing:
                cur.execute(
                    """UPDATE ground_truth SET
                         country_name = ?, country_group = ?,
                         cross_year_id = ?, dimension = ?, indicator = ?,
                         response = ?, decision = ?,
                         awarded_score = ?, max_score = ?,
                         explanation = ?, loaded_at = datetime('now')
                       WHERE id = ?""",
                    (*p[2:], existing[0]),
                )
                updated += 1
            else:
                cur.execute(
                    """INSERT INTO ground_truth
                       (question_id, country_code, country_name, country_group,
                        cross_year_id, dimension, indicator,
                        response, decision, awarded_score, max_score,
                        explanation)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    p,
                )
                inserted += 1
        conn.commit()

    print(
        f"Loaded merged_responses: inserted {inserted}, updated {updated}. "
        f"Table ground_truth now has "
        f"{inserted + updated} rows for cycle 2025."
    )


if __name__ == "__main__":
    main()
